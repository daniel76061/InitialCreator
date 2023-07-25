from openpyxl import load_workbook

def generate_initials(name, surname, existing_initials):
    # Generate initials using the first letters of the first name and surname
    initials = name[0].upper() + surname[0].upper()

    # Check if 2-letter initials already exist, generate 3-letter initials if needed
    index = 1
    while initials in existing_initials:
        if len(surname) > index:
            initials = name[0].upper() + surname[0].upper() + surname[index].upper()
            index += 1
        else:
            break

    # Add the generated initials to the set of existing initials
    existing_initials.add(initials)
    return initials

if __name__ == "__main__":
    firstname = input("Please insert firstname: ").capitalize()
    lastname = input("Please insert last name: ").capitalize()

    # Load the existing Excel workbook or create a new one if not found
    try:
        wb = load_workbook("Spreadsheet.xlsx")
    except FileNotFoundError:
        wb = load_workbook()
        wb.create_sheet("Sheet1")
    ws = wb.active

    # Get the 'D' column values for existing initials in the worksheet
    existing_initials = set(cell.value for cell in ws['D'][1:])

    # Generate the initials for the new user
    initials = generate_initials(firstname, lastname, existing_initials)

    # Append the new user data to the worksheet
    fullname = f"{firstname} {lastname}"
    email = f"{firstname.lower()}.{lastname.lower()}@address.com"
    username = f"Domain\\{firstname.lower()}.{lastname.lower()}"

    ws.append([fullname, email, username, initials.upper()])

    # Save the updated workbook
    wb.save("Spreadsheet.xlsx")

    # Print the new user data in one line
    print(f"New user data: Full Name: {fullname}, Email: {email}, Username: {username}, Initials: {initials.upper()}")
